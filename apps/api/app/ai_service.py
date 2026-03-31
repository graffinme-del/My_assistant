from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy.orm import Session

from .config import settings
from .models import Case, CaseEvent, CaseTag, Task


def _chat_completions_url() -> str:
    base = settings.openai_base_url.rstrip("/")
    return f"{base}/chat/completions"


def _llm_headers() -> dict[str, str]:
    h: dict[str, str] = {
        "Authorization": f"Bearer {settings.openai_api_key}",
        "Content-Type": "application/json",
    }
    if settings.llm_http_referer.strip():
        h["HTTP-Referer"] = settings.llm_http_referer.strip()
    if settings.llm_app_title.strip():
        h["X-Title"] = settings.llm_app_title.strip()
    return h


async def _llm_chat(user_content: str, *, timeout: float = 45.0) -> str:
    payload = {
        "model": settings.openai_model,
        "messages": [{"role": "user", "content": user_content}],
    }
    async with httpx.AsyncClient(timeout=timeout) as client:
        resp = await client.post(_chat_completions_url(), headers=_llm_headers(), json=payload)
    try:
        resp.raise_for_status()
    except httpx.HTTPStatusError as e:
        code = e.response.status_code
        if code == 401:
            raise ValueError(
                "OpenAI/OpenRouter вернул 401: ключ неверный, не тот для этого URL, или истёк. "
                "На сервере в /opt/my_assistant/.env задайте OPENAI_API_KEY (sk-… с platform.openai.com). "
                "Если ключ от OpenRouter — OPENAI_BASE_URL=https://openrouter.ai/api/v1 и модель вида openai/gpt-4o-mini."
            ) from e
        if code == 429:
            raise ValueError(
                "Провайдер вернул 429 (лимит или нет оплаты). Проверьте биллинг и квоты."
            ) from e
        raise
    data = resp.json()
    choices = data.get("choices") or []
    if not choices:
        return ""
    msg = (choices[0].get("message") or {}).get("content") or ""
    return str(msg).strip()


def _extract_date(text: str) -> date | None:
    match = re.search(r"(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?", text)
    if not match:
        return None
    day = int(match.group(1))
    month = int(match.group(2))
    year_raw = match.group(3)
    year = int(year_raw) if year_raw else date.today().year
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def build_case_summary(case: Case, events: list[CaseEvent], tasks: list[Task]) -> str:
    open_tasks = [t for t in tasks if t.status != "done"]
    latest_events = sorted(events, key=lambda e: e.created_at, reverse=True)[:3]
    events_block = "\n".join([f"- {e.event_type}: {e.body[:160]}" for e in latest_events]) or "- нет"
    tasks_block = "\n".join([f"- {t.title} ({t.priority})" for t in open_tasks[:5]]) or "- нет"
    tags_block = ", ".join(sorted({t.value for t in getattr(case, "tags", [])})) or "нет"
    return (
        f"Дело: {case.title}\n"
        f"Статус: {case.status}, стадия: {case.stage}\n"
        f"Теги/алиасы: {tags_block}\n"
        f"Следующее заседание: {case.next_hearing_date or 'не указано'}\n"
        f"Последние события:\n{events_block}\n"
        f"Открытые задачи:\n{tasks_block}"
    )


async def llm_assistant_chat_reply(user_message: str, case: Case) -> str:
    """Ответ в стиле чата (не юридическая консультация — организация и краткие пояснения)."""
    if not settings.openai_api_key.strip():
        return (
            "Сообщение сохранено. Чтобы я отвечал как ассистент на базе модели, "
            "нужен ключ API (OPENAI_API_KEY; для OpenRouter также OPENAI_BASE_URL)."
        )

    user_content = (
        "Ты личный помощник по организации материалов по судебным делам. "
        "Отвечай по-русски, кратко и по существу (обычно 2–6 предложений). "
        "Не выдавай юридических консультаций: не оценивай шансы, не подсказывай стратегию и не указывай, "
        "какие нормы «должны» применяться. Можно помочь структурировать мысли, напомнить о сроках и фактах из контекста. "
        "Если пользователь уже пишет, что архив/файлы загружены, не проси загрузить их повторно. "
        "Если данных не хватает, задай только один самый полезный уточняющий вопрос. "
        "Если речь о разборе документов по делам, предлагай конкретное следующее действие, а не общие рассуждения.\n\n"
        f"Контекст: дело «{case.title}», номер {case.case_number}, статус {case.status}.\n\n"
        f"Сообщение пользователя:\n{user_message}"
    )
    text = await _llm_chat(user_content, timeout=90.0)
    return text or "Не удалось получить текст ответа от модели."


async def llm_summary(prompt: str) -> str:
    if not settings.openai_api_key:
        return "LLM ключ не настроен. Возвращаю локальную сводку без внешнего ИИ."

    user_content = (
        "Ты личный помощник по судебным делам. Дай короткую сводку в 5-8 строках. "
        "Только факты из входных данных.\n\n"
        f"{prompt}"
    )
    text = await _llm_chat(user_content, timeout=60.0)
    return text or "Не удалось получить ответ модели."


def parse_hearing_note(db: Session, case: Case, text: str) -> tuple[CaseEvent, list[Task], date | None]:
    extracted_date = _extract_date(text)
    if extracted_date:
        case.next_hearing_date = extracted_date

    event = CaseEvent(case_id=case.id, event_type="hearing_note", body=text)
    db.add(event)

    tasks: list[Task] = []
    for line in text.splitlines():
        normalized = line.strip("- ").strip()
        if not normalized:
            continue
        if any(token in normalized.lower() for token in ["подготов", "приобщ", "направ", "предостав"]):
            task = Task(
                case_id=case.id,
                title=normalized[:200],
                description="Создано автоматически из сообщения после заседания",
                priority="high",
                status="open",
            )
            db.add(task)
            tasks.append(task)

    db.commit()
    db.refresh(event)
    for task in tasks:
        db.refresh(task)

    return event, tasks, extracted_date


def extract_document_text(file_path: Path, filename: str) -> str:
    suffix = filename.lower().split(".")[-1] if "." in filename else ""
    if suffix in {"txt", "md"}:
        return file_path.read_text(encoding="utf-8", errors="ignore")
    if suffix == "pdf":
        reader = PdfReader(str(file_path))
        chunks: list[str] = []
        for page in reader.pages[:30]:
            chunks.append(page.extract_text() or "")
        return "\n".join(chunks).strip()
    if suffix in {"csv", "log"}:
        return file_path.read_text(encoding="utf-8", errors="ignore")
    if suffix == "xlsx":
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        chunks: list[str] = []
        for ws in wb.worksheets[:10]:
            rows_seen = 0
            chunks.append(f"Лист: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if not values:
                    continue
                chunks.append(" | ".join(values[:20]))
                rows_seen += 1
                if rows_seen >= 200:
                    break
        return "\n".join(chunks).strip()
    return ""


def classify_document(filename: str, text: str) -> tuple[str, float]:
    probe = f"{filename}\n{text}".lower()
    mapping = [
        ("court_act", ["определение", "решение суда", "постановление"]),
        ("power_of_attorney", ["доверенность", "power of attorney"]),
        ("claim", ["исковое заявление", "иск", "claim"]),
        ("review", ["отзыв", "возражения"]),
        ("complaint", ["жалоба", "апелляция", "кассация"]),
        ("evidence", ["cmr", "доказательств", "приложени"]),
    ]
    for category, keys in mapping:
        if any(key in probe for key in keys):
            return category, 0.85
    return "other", 0.55


def _normalize(value: str) -> str:
    return re.sub(r"[^a-zа-я0-9]", "", value.lower())


def _tokenize_tag_values(raw: str) -> list[str]:
    quoted = [m.strip() for m in re.findall(r'"([^"]+)"|«([^»]+)»', raw) for m in m if m.strip()]
    cleaned = re.sub(r'"[^"]+"|«[^»]+»', ",", raw)
    parts = re.split(r"[,\n;•|]+", cleaned)
    values = quoted + [p.strip(" -\t\r\n") for p in parts]
    deduped: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not value:
            continue
        key = _normalize(value)
        if not key or key in seen:
            continue
        seen.add(key)
        deduped.append(value)
    return deduped


def find_case_by_hint(cases: list[Case], hint: str) -> Case | None:
    norm_hint = _normalize(hint)
    if not norm_hint:
        return None

    best: Case | None = None
    best_score = 0.0
    for case in cases:
        score = 0.0
        title_norm = _normalize(case.title)
        if title_norm and (title_norm in norm_hint or norm_hint in title_norm):
            score += 0.8
        if case.case_number and _normalize(case.case_number) in norm_hint:
            score += 0.95
        for tag in getattr(case, "tags", []):
            tag_norm = _normalize(tag.value)
            if not tag_norm:
                continue
            if tag_norm in norm_hint or norm_hint in tag_norm:
                score += 0.7 if tag.kind == "alias" else 0.45
        if score > best_score:
            best_score = score
            best = case
    return best if best_score >= 0.65 else None


def looks_like_case_tag_update(text: str) -> bool:
    t = text.lower()
    return (
        any(marker in t for marker in ["теги", "теги", "ключевые слова", "алиасы", "алиас", "синоним"])
        and any(marker in t for marker in ["дело", "папк", "банкрот"])
    )


def parse_case_tag_update(text: str) -> dict[str, Any] | None:
    if not looks_like_case_tag_update(text):
        return None

    normalized = " ".join(text.replace("\r", "\n").split())
    tag_match = re.search(
        r"(?:теги|теги|ключевые слова|алиасы|алиас|синонимы?)\s*[:\-]\s*(.+)$",
        normalized,
        flags=re.IGNORECASE,
    )
    if not tag_match:
        return None

    tags_raw = tag_match.group(1).strip()
    before_tags = normalized[: tag_match.start()].strip(" .:-")
    case_match = re.search(
        r"(?:для|по)\s+(?:папк[еи]|дел[ауо]?|банкротств[ауо]?)\s+(.+)$|дело\s+(.+)$",
        before_tags,
        flags=re.IGNORECASE,
    )
    case_hint = ""
    if case_match:
        case_hint = next((g for g in case_match.groups() if g), "").strip(" .:-")
    if not case_hint:
        return None

    aliases = _tokenize_tag_values(re.sub(r"\s+или\s+", ", ", case_hint, flags=re.IGNORECASE))
    tags = _tokenize_tag_values(tags_raw)
    if not tags:
        return None
    return {
        "case_hint": case_hint,
        "title_candidate": aliases[0] if aliases else case_hint,
        "aliases": aliases,
        "tags": tags,
    }


async def llm_parse_case_tag_update(text: str, cases: list[Case]) -> dict[str, Any] | None:
    if not settings.openai_api_key.strip() or not looks_like_case_tag_update(text):
        return None
    prompt = (
        "Извлеки из сообщения пользователя настройку тегов для судебного дела.\n"
        "Верни JSON строго такого вида:\n"
        '{"case_hint":"...","title_candidate":"...","aliases":["..."],"tags":["..."]}\n'
        "Если это не запрос на сохранение тегов по делу, верни {}.\n"
        "aliases — варианты названия дела. tags — ключевые слова для автопривязки документов.\n"
        f"Уже существующие дела: {[c.title for c in cases[:30]]}\n\n"
        f"Сообщение:\n{text}"
    )
    raw = await _llm_chat(prompt, timeout=45.0)
    try:
        parsed = json.loads(raw)
    except Exception:
        return None
    if not isinstance(parsed, dict) or not parsed.get("tags") or not parsed.get("case_hint"):
        return None
    parsed["aliases"] = _tokenize_tag_values(", ".join(parsed.get("aliases") or []))
    parsed["tags"] = _tokenize_tag_values(", ".join(parsed.get("tags") or []))
    parsed["title_candidate"] = str(parsed.get("title_candidate") or parsed["case_hint"]).strip()
    parsed["case_hint"] = str(parsed["case_hint"]).strip()
    return parsed if parsed["tags"] and parsed["case_hint"] else None


def match_case(db: Session, filename: str, text: str, preferred_case_id: int | None = None) -> tuple[Case | None, float]:
    if preferred_case_id:
        case = db.query(Case).filter(Case.id == preferred_case_id).first()
        if case:
            return case, 0.9

    cases = db.query(Case).all()
    if not cases:
        return None, 0.0

    corpus = _normalize(f"{filename}\n{text}")[:30000]
    for case in cases:
        if _normalize(case.case_number) in corpus and len(_normalize(case.case_number)) > 4:
            return case, 0.95

    best: Case | None = None
    best_score = 0.0
    for case in cases:
        score = 0.0
        title_token = _normalize(case.title)
        court_token = _normalize(case.court_name)
        if title_token and title_token[:12] in corpus:
            score += 0.35
        if court_token and court_token[:12] in corpus:
            score += 0.25
        for tag in getattr(case, "tags", []):
            tag_token = _normalize(tag.value)
            if not tag_token or len(tag_token) < 3:
                continue
            if tag_token in corpus:
                score += 0.7 if tag.kind == "alias" else 0.45
        if score > best_score:
            best_score = score
            best = case
    if best:
        return best, max(0.6, best_score)
    return cases[0], 0.45


def extract_case_number(text: str) -> str | None:
    # Examples:
    # - А40-12345/2026
    # - 2-123/2026
    # Keep it intentionally permissive for messy Telegram copies.
    patterns = [
        r"([АA]\d{1,4}\s*-\s*\d{1,7}\s*/\s*\d{2,4})",
        r"(\d{1,2}\s*-\s*\d{1,7}\s*/\s*\d{2,4})",
        r"(\d{1,4}\s*\/\s*\d{2,4})",
    ]
    for p in patterns:
        m = re.search(p, text, flags=re.IGNORECASE)
        if m:
            raw = m.group(1) if m.lastindex else m.group(0)
            raw = raw.replace(" ", "").replace("\n", "")
            raw = raw.replace("\\", "")
            return raw
    return None


def looks_like_hearing_note(text: str) -> bool:
    t = text.lower()
    keywords = [
        "отлож",
        "заседан",
        "судья",
        "попросил",
        "письменно",
        "приобщ",
        "доказательств",
        "экземпляр",
        "сопостав",
        "залуч",
    ]
    if any(k in t for k in keywords):
        return True
    # Date-only hint.
    return bool(re.search(r"\d{1,2}\.\d{1,2}", text))


async def llm_document_routing(
    *,
    filename: str,
    text: str,
    available_case_numbers: list[str],
) -> dict[str, Any] | None:
    if not settings.openai_api_key:
        return None

    text_sample = text[:8000]
    prompt = (
        "Ты помощник для сортировки судебных документов.\n"
        "Верни JSON строго в формате:\n"
        '{"category":"...","case_number":"...","confidence":0.0,"short_note":"..."}\n'
        "category выбирай из: court_act, power_of_attorney, claim, review, complaint, evidence, correspondence, other.\n"
        "case_number выбери из списка доступных номеров дел, либо пустую строку если не уверен.\n"
        "confidence от 0 до 1.\n\n"
        f"Доступные номера дел: {available_case_numbers}\n"
        f"Имя файла: {filename}\n"
        f"Текст:\n{text_sample}"
    )
    raw = await _llm_chat(prompt, timeout=60.0)
    try:
        import json

        return json.loads(raw)
    except Exception:
        return None
